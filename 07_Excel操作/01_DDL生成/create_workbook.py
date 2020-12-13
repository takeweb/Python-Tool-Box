from openpyxl import Workbook

wb = Workbook()
ws0 = wb.active
ws0.title = 'ワークシート0'
ws1 = wb.create_sheet('ワークシート1')
ws2 = wb.create_sheet('ワークシート2')

wb.active = 2
wb.save('new_workbook_2.xlsx')
