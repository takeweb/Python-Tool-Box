#!/usr/bin/env python
# -*- coding: utf-8 -*-

import sys, argparse, datetime, os, re, shutil
from openpyxl import load_workbook

def createDDL(filename, out_dir_name):
    # DDL生成
    ROW_HEADER_MIN = 4
    ROW_DETAIL_MIN = 13
    row_detail_max = 0
    sheet_list = []

    # テーブル定義書Excelファイル読み込み
    wb = load_workbook(filename, read_only=True, data_only=True)

    ws = wb["テーブル一覧"]
    for i, row in enumerate(ws):
        if i >= 2 and row[3].value is not None:
           if row[3].value == "〇":
               sheet_list.append(row[1].value)
    
    # シート名ループ
#   print(sheet_list)
    for sheet_name in sheet_list:
        #print(sheet_name)
        col_info_list = []
        idx_info_list = []
        if sheet_name == "テーブル一覧":
            continue
        ws = wb[sheet_name]
        #print(ws)
        end_flg = False
        idx_min_row = 0
        idx_max_row = 0
        table_name = ""
        table_remarks = ""

        # 各情報の開始位置、終了位置を把握
#       print("各情報の開始位置、終了位置を把握開始!!")
        for i, row in enumerate(ws):
            #print(row[0].value)
            if row[0].value == "インデックス情報":
                #print("インデックス情報")
                row_detail_max = i
                idx_min_row = i + 2
                #print(row_detail_max)
                #print(idx_min_row)
            elif row[0].value == "制約情報":
                #print("制約情報")
                idx_max_row = i
                #print(idx_max_row)
                break
#       print("各情報の開始位置、終了位置を把握完了!!")

        # インデックス情報
        SQL = ""
        for i, row in enumerate(ws):
            if i == ROW_HEADER_MIN + 1:
                table_name = str(row[2].value).rstrip()

            if i >= idx_min_row and i < idx_max_row and row[1].value is not None:
                idx_name = ""
                idx_cols = ""
                idx_name = str(row[1].value)
                idx_cols = str(row[2].value)
                #print(idx_name)
                #print(idx_cols)

                if re.compile("^pk.").search(idx_name):
                    continue
                elif re.compile("^uq.").search(idx_name):
                    SQL = "ALTER TABLE ONLY " + table_name + " ADD CONSTRAINT "+ idx_name + " UNIQUE (" + idx_cols + ");"
                else:
                    SQL = "CREATE INDEX " + idx_name + " ON " + table_name + " USING btree (" + idx_cols + ");"
                idx_info_list.append(SQL)

        # 行ループ
        for i, row in enumerate(ws):
            col_info = {}
            idx_col = ""

            if i == ROW_HEADER_MIN:
                table_remarks = str(row[2].value).rstrip()
            elif i == ROW_HEADER_MIN + 1:
                table_name = str(row[2].value).rstrip()
            elif i == row_detail_max:
                break
            
            # 列ループ
            for j, col in enumerate(row): 
                if col.value is not None:
                    if col.data_type == "d":
                        value = col.value
                    else:
                        value = str(col.value).rstrip()

                    if i >= ROW_DETAIL_MIN:
                        if j == 1:
                            col_info["remarks"] = value
                        elif j == 2:
                            col_info["name"] = value
                            idx_col = value
                        elif j == 3:
                            col_info["data_type"] = value.upper()
                        elif j == 4 and value != "":
                            col_info["not_null"] = " NOT NULL"
                            if value.find("PK") > -1:
                                col_info["pk"] = "  , CONSTRAINT pk_" + table_name + "_" + idx_col + " PRIMARY KEY (" + idx_col + ")"
                        elif j == 5 and value != "":
                            col_info["default"] = " DEFAULT " + value
                        #print(f'data:{value}')
            if len(col_info) > 0:
                #print(col_info)
                col_info_list.append(col_info)
            if end_flg:
                break
        #print(table_name)
        #print(table_remarks)
        #print(col_info_list)
        writeFile(table_name, table_remarks, col_info_list, idx_info_list, out_dir_name)
    wb.close()

def writeFile(table_name, table_remarks, col_info_list, idx_info_list, out_dir_name):
    col = ""
    pk = ""

    # 出力ファイル名生成
    out_file_name = get_file_name(out_dir_name, table_name)

    # テキストファイルへ出力
    with open(out_file_name, "w", encoding = "utf_8", newline='\n') as out_f:
        out_f.write("-- " + table_remarks + "\n")
        out_f.write("DROP TABLE IF EXISTS " + table_name + " CASCADE;\n")
        out_f.write("CREATE TABLE " + table_name + "(\n")
        for i, col_info in enumerate(col_info_list):
            col = ""
            if i == 0:
                col += "    "
            else:
                col += "  , "
            col += col_info["name"] + " " + col_info["data_type"]
            if "not_null" in col_info:
                col += col_info["not_null"]
            if "default" in col_info:
                col += col_info["default"]
            col += "\n"
            out_f.write(col)
            if "pk" in col_info:
                pk = col_info["pk"]
        out_f.write(pk + "\n")
        out_f.write(");\n\n")

        # コメント出力
        out_f.write("COMMENT ON TABLE " + table_name + " IS '" + table_remarks + "';\n")
        for i, col_info in enumerate(col_info_list):
            col = "COMMENT ON COLUMN " 
            col += table_name + "." + col_info["name"] + " IS '" + col_info["remarks"] + "';\n"
            out_f.write(col)
        out_f.write("\n")

        # インデックス出力(PK除く)
        for idx in idx_info_list:
            out_f.write(idx + "\n")

def get_file_name(dir_name, table_name):
    file_name = os.path.join(dir_name, table_name + ".sql")
    if os.path.exists(file_name):
        os.remove(file_name)

    return file_name

def get_dir_name():
    # 出力ファイル保存先
    save_dir_base = "C:\\Users\\oishi\\Documents\\99_WORK\\"

    # 日付取得・分解
    dt_now = datetime.datetime.now()
    year = str(dt_now.year)
    month = str(dt_now.month).zfill(2)
    day = str(dt_now.day).zfill(2)

    save_list = [year, year + month, year + month + day, "DDL"]
    dir_name = os.path.join(save_dir_base, *save_list)

    # 出力先ディレクトリの存在チェック
    if not os.path.exists(dir_name):
        os.makedirs(dir_name)
    return dir_name

def outTimestamp(filename, dt_out):
    # DDL最終出力日時を記入

    # テーブル定義書Excelファイル読み込み
    wb = load_workbook(filename, read_only=False, data_only=True)

    ws = wb["テーブル一覧"]
    for i, row in enumerate(ws):
        if i >= 2 and row[3].value is not None:
           if row[3].value == "〇":
               ws.cell(row=i + 1, column=5).value = dt_out
    wb.save(filename)

if __name__ == '__main__':
    try:
        # コマンドライン引数設定
        parser = argparse.ArgumentParser()
        parser.add_argument('-f', '--file')

        # コマンドライン引数受け取り
        args = parser.parse_args()
        input_file = args.file

        # 出力先ディレクトリ準備
        out_dir_name = get_dir_name()

        # DDL生成関数呼び出し
        createDDL(input_file, out_dir_name)

        # DDL最終出力日時を記入
        dt_now = datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')
        outTimestamp(input_file, dt_now)
        print("Created Files!! at " + dt_now)

    except Exception as e:
        print(e)

